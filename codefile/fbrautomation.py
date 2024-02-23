#!/usr/bin/env python
# coding: utf-8

# In[17]:


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
driver = webdriver.Chrome()

# Replace 'file.xlsx' and 'Sheet1' with your actual file path and sheet name
excel_file_path = 'file.xlsx'
sheet_name = 'Sheet1'

# Read the Excel file
excel_data = pd.read_excel(excel_file_path, sheet_name=sheet_name,header=None)

# Open the Excel workbook
wb = openpyxl.load_workbook(excel_file_path)

# Access the active sheet (assuming it's the first sheet)
ws = wb.active

# Create an empty list to store the formatted strings
formatted_strings = []

# Iterate through each value in the column and format it
for value in excel_data.iloc[:, 0]:
    # Check if the string is already in the desired format
    if len(value) == 24 and value.count('-') == 3:
        formatted_string = value
    else:
        # Apply formatting if not in the desired format
        formatted_string = '{}-{}-{}-{}'.format(value[:2], value[2:6], value[6:11], value[11:])
    
    # Append the formatted string to the list
    formatted_strings.append(formatted_string)
    #print(formatted_string)
    
#automation starts here
def automate_navigation(main_window):
    # Set the URL of the website
    website_url = "https://e.fbr.gov.pk/"

    # Set up the Chrome driver (make sure chromedriver is in your PATH or specify its path)
    #driver = webdriver.Chrome()

    try:
        # Open the specified URL
        driver.get(website_url)

        # Find the "verification" button element
        verification_button_xpath = '//*[@id="ctl00_MainDetailHeader1_mnTopNavigationn3"]/table/tbody/tr/td[1]/a'
        verification_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, verification_button_xpath))
        )

        # Create an ActionChains object and perform a mouseover action on the "verification" button
        actions = ActionChains(driver)
        actions.move_to_element(verification_button).perform()

        # Wait for the second element to be clickable
        second_element_xpath = '//*[@id="ctl00_MainDetailHeader1_mnTopNavigationn18"]/td/table/tbody/tr/td/a'
        second_element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, second_element_xpath))
        )

        # Click the second element
        second_element.click()
        
        
                # Locate the dropdown element
        dropdown_xpath = '//*[@id="ctl00_ContentPlaceHolder1_ddlSearchParam"]'  # Replace with the actual XPath or other locator
        dropdown = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, dropdown_xpath))
        )

        # Create an ActionChains object and click on the dropdown to open it
        actions = ActionChains(driver)
        actions.click(dropdown).perform()
        
        dropdown_xpath = '//*[@id="ctl00_ContentPlaceHolder1_ddlSearchParam"]'  # Replace with the actual XPath or other locator
        dropdown = WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.XPATH, dropdown_xpath))
        )

        # Use the Select class to interact with the dropdown
        dropdown_select = Select(dropdown)

        # Select the desired option by index, value, or visible text
        dropdown_select.select_by_index(1)
        for formatted_string in formatted_strings:
            # Locate the input field by ID
            
            input_field_id = 'ctl00_ContentPlaceHolder1_txtEDNBarCode'  # Replace with the ID
            input_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, input_field_id))
            )
            input_field.click()
            input_field.clear()
            WebDriverWait(driver, 2)
            # Enter text into the input field
            #input_field.send_keys("IT-2022-09120-1011408761")  # Replace with the desired text
            input_field.send_keys(formatted_string) # Replace with the desired text
            print(formatted_string)
            search_button_id = 'ctl00_ContentPlaceHolder1_btnSearch'  # Replace with the actual ID
            search_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, search_button_id))
            )
        
            # Click the search button
            search_button.click()
        
             # Wait for the pop-up window to appear
            popup_window = WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))

            # Switch to the pop-up window
            windows = driver.window_handles
            popup_window_handle = next(handle for handle in windows if handle not in main_window)
            driver.switch_to.window(popup_window_handle)

            # Locate the link in the pop-up window by XPath
            link_xpath = '//*[@id="lnkBtnPrint"]'  # Replace with the actual XPath
            link = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, link_xpath))
            )

            # Click the link in the pop-up window
            link.click()
        
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="lnkBtnPrint"]'))
            )
            driver.close()
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == formatted_string:
                        # Highlight the cell with a yellow fill
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            WebDriverWait(driver, 1)
            print("Download completed.")
            driver.switch_to.window(main_window)
            # Optional: Add a delay or wait for user input to keep the browser window open"""
        #input("Press Enter to close the browser...")
        wb.save(excel_file_path)
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        # Print the full traceback for better debugging
        import traceback
        traceback.print_exc()

    finally:
        # Quit the browser at the end of the script
        driver.quit()

if __name__ == "__main__":
    
    main_window = driver.window_handles[0]
    automate_navigation(main_window)

